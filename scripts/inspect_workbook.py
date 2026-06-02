#!/usr/bin/env python3
"""Inspect an OB survey scale workbook and produce structured review inputs.

This script is intentionally conservative. It extracts workbook structure,
variable blocks, questionnaire items, placeholders, reverse-coded markers, and
basic consistency issues. The research judgment and final review should still be
done by the agent using the skill references.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PLACEHOLDER_RE = re.compile(
    r"(XX|xx|x月x日|X月X日|某某|姓名|name of school|name of organization|____+|【.*?】)"
)
REVERSE_RE = re.compile(r"\b(reverse|reverse[- ]?coded|reversed|反向|反向计分)\b", re.I)
TIME_RE = re.compile(r"^T\d+$", re.I)
INSTRUCTION_CONTAMINATION_RE = re.compile(
    r"(这里关注的是|而不是|定义为|指的是|以下陈述描述的是您对|以下描述涉及您对)"
)
REFERENT_AMBIGUITY_RE = re.compile(r"(本部门|部门|我们|团队成员|他人)")
DOUBLE_BARRELED_RE = re.compile(r"(和|与|以及|并|同时|或|、)")


@dataclass
class VariableRecord:
    sheet: str
    row: int
    english_name: str
    chinese_name: str
    item_count: str
    respondent: str
    wave: str
    questionnaire_id: str


@dataclass
class ItemRecord:
    sheet: str
    row: int
    wave: str
    respondent: str
    variable: str
    source_original: str
    current_chinese: str
    source_reference: str
    note: str
    block_start_row: int
    block_instruction: str


@dataclass
class IssueRecord:
    issue_id: str
    priority: str
    variable_name: str
    location: str
    issue_type: str
    issue_summary: str
    current_text: str
    suggested_text: str
    rationale: str
    reviewer_action: str
    status: str = "待处理"


def clean(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).replace("\r", "\n").strip()


def norm(value: str) -> str:
    return re.sub(r"\s+", "", value or "").lower()


def read_workbook(path: Path) -> dict[str, pd.DataFrame]:
    return pd.read_excel(path, sheet_name=None, header=None, dtype=object)


def sheet_sample(df: pd.DataFrame, max_rows: int = 12, max_cols: int = 8) -> list[list[str]]:
    sample = df.iloc[: min(max_rows, len(df)), : min(max_cols, df.shape[1])]
    return [[clean(v) for v in row] for row in sample.to_numpy().tolist()]


def find_header_row(df: pd.DataFrame, required_terms: list[str]) -> int | None:
    for idx, row in df.iterrows():
        row_text = " ".join(clean(v) for v in row.tolist())
        if all(term in row_text for term in required_terms):
            return int(idx)
    return None


def map_headers(df: pd.DataFrame, header_row: int) -> dict[str, int]:
    headers = {}
    for col, value in enumerate(df.iloc[header_row].tolist()):
        text = clean(value)
        if text:
            headers[text] = col
    return headers


def col_for(headers: dict[str, int], candidates: list[str]) -> int | None:
    for cand in candidates:
        for header, idx in headers.items():
            if cand.lower() in header.lower():
                return idx
    return None


def extract_variables(sheets: dict[str, pd.DataFrame]) -> list[VariableRecord]:
    records: list[VariableRecord] = []
    for sheet_name, df in sheets.items():
        header_row = find_header_row(df, ["变量名称", "量表条目数"])
        if header_row is None:
            continue
        headers = map_headers(df, header_row)
        c_eng = col_for(headers, ["变量名称"])
        c_cn = col_for(headers, ["变量名称-中文", "中文"])
        c_count = col_for(headers, ["量表条目数", "条目数"])
        c_resp = col_for(headers, ["填写人", "评价者"])
        c_wave = col_for(headers, ["Time", "时间", "Wave"])
        c_qid = col_for(headers, ["问卷编号"])
        if c_cn is None:
            continue
        for row_idx in range(header_row + 1, len(df)):
            row = df.iloc[row_idx]
            first_text = " ".join(clean(v) for v in row.tolist())
            if "sum item num" in first_text.lower() or first_text.startswith("SUM"):
                break
            cn = clean(row.iloc[c_cn])
            eng = clean(row.iloc[c_eng]) if c_eng is not None else ""
            if not cn and not eng:
                continue
            records.append(
                VariableRecord(
                    sheet=sheet_name,
                    row=row_idx + 1,
                    english_name=eng,
                    chinese_name=cn,
                    item_count=clean(row.iloc[c_count]) if c_count is not None else "",
                    respondent=clean(row.iloc[c_resp]) if c_resp is not None else "",
                    wave=clean(row.iloc[c_wave]) if c_wave is not None else "",
                    questionnaire_id=clean(row.iloc[c_qid]) if c_qid is not None else "",
                )
            )
    return records


def is_questionnaire_header(row: list[str]) -> bool:
    row_text = " ".join(row)
    return "Measures" in row_text or "评价者" in row_text or "Source Reference" in row_text


def extract_items(sheets: dict[str, pd.DataFrame]) -> list[ItemRecord]:
    items: list[ItemRecord] = []
    for sheet_name, raw in sheets.items():
        header_row = None
        for idx, row in raw.iterrows():
            vals = [clean(v) for v in row.tolist()]
            if is_questionnaire_header(vals):
                header_row = int(idx)
                break
        if header_row is None:
            continue

        header = [clean(v) for v in raw.iloc[header_row].tolist()]
        col_map = {h: i for i, h in enumerate(header) if h}
        c_eval = col_for(col_map, ["评价者", "rater", "respondent"])
        c_var = col_for(col_map, ["Variables", "变量"])
        c_orig = col_for(col_map, ["original", "原始"])
        c_cn = col_for(col_map, ["Chinese", "中文", "adapted"])
        c_source = col_for(col_map, ["Source", "来源"])
        c_note = col_for(col_map, ["备注", "note"])
        if c_var is None or c_cn is None:
            continue

        current_wave = ""
        current_respondent = ""
        current_variable = ""
        current_source = ""
        current_note = ""
        current_instruction = ""
        block_start_row = 0

        for row_idx in range(header_row + 1, len(raw)):
            row = raw.iloc[row_idx]
            vals = [clean(v) for v in row.tolist()]
            respondent = clean(row.iloc[c_eval]) if c_eval is not None and c_eval < len(row) else ""
            variable = clean(row.iloc[c_var]) if c_var < len(row) else ""
            original = clean(row.iloc[c_orig]) if c_orig is not None and c_orig < len(row) else ""
            chinese = clean(row.iloc[c_cn]) if c_cn < len(row) else ""
            source = clean(row.iloc[c_source]) if c_source is not None and c_source < len(row) else ""
            note = clean(row.iloc[c_note]) if c_note is not None and c_note < len(row) else ""

            if variable and TIME_RE.match(variable):
                current_wave = variable.upper()
                continue

            # A block row usually has respondent/variable/source/instruction text.
            if variable and (respondent or source or (chinese and not original)):
                current_respondent = respondent or current_respondent
                current_variable = variable
                current_source = source
                current_note = note
                current_instruction = chinese
                block_start_row = row_idx + 1
                if original:
                    items.append(
                        ItemRecord(
                            sheet=sheet_name,
                            row=row_idx + 1,
                            wave=current_wave,
                            respondent=current_respondent,
                            variable=current_variable,
                            source_original=original,
                            current_chinese=chinese,
                            source_reference=current_source,
                            note=current_note,
                            block_start_row=block_start_row,
                            block_instruction=current_instruction,
                        )
                    )
                continue

            if original or chinese:
                items.append(
                    ItemRecord(
                        sheet=sheet_name,
                        row=row_idx + 1,
                        wave=current_wave,
                        respondent=current_respondent,
                        variable=current_variable,
                        source_original=original,
                        current_chinese=chinese,
                        source_reference=current_source,
                        note=current_note,
                        block_start_row=block_start_row,
                        block_instruction=current_instruction,
                    )
                )
    return items


def variable_key(text: str) -> str:
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[（）()]", "", text)
    return text.lower()


def respondent_alias(value: str) -> str:
    v = norm(value)
    if v in {"下属", "员工", "employee", "subordinate"}:
        return "员工"
    if v in {"领导", "主管", "leader", "supervisor"}:
        return "领导"
    return value.strip()


def build_issues(variables: list[VariableRecord], items: list[ItemRecord]) -> list[IssueRecord]:
    issues: list[IssueRecord] = []
    seen: set[tuple[str, str, str, str]] = set()
    flagged_instructions: set[tuple[str, int, str]] = set()

    def add(
        priority: str,
        variable_name: str,
        location: str,
        issue_type: str,
        current: str,
        suggested: str,
        rationale: str,
        action: str,
    ) -> None:
        key = (priority, variable_name, location, issue_type, current[:300])
        if key in seen:
            return
        seen.add(key)
        issues.append(
            IssueRecord(
                issue_id=f"I{len(issues) + 1:03d}",
                priority=priority,
                variable_name=variable_name,
                location=location,
                issue_type=issue_type,
                issue_summary=summarize_issue(issue_type, variable_name),
                current_text=current,
                suggested_text=suggested,
                rationale=rationale,
                reviewer_action=action,
            )
        )

    for item in items:
        loc = f"{item.sheet}!row {item.row}"
        variable_name = item.variable.split("\n")[0] if item.variable else ""
        launch_text = "\n".join([item.variable, item.current_chinese, item.note, item.block_instruction])
        if PLACEHOLDER_RE.search(launch_text):
            if PLACEHOLDER_RE.search(item.block_instruction):
                placeholder_loc = f"{item.sheet}!row {item.block_start_row}"
                placeholder_text = "\n".join([item.variable, item.block_instruction])
            else:
                placeholder_loc = loc
                placeholder_text = launch_text
            add("P0", variable_name, placeholder_loc, "placeholder", placeholder_text[:500], "替换所有正式发放前占位符", "问卷中仍有占位符或内部模板文本。", "RA 可直接处理")
        elif PLACEHOLDER_RE.search(item.source_original):
            add("P3", variable_name, loc, "source_placeholder", item.source_original, "通常无需修改原英文；确认中文改编已替换具体对象", "原英文成熟量表含模板占位符，这本身不是发放阻断，但审查时要确认中文已完成情境替换。", "仅提示")
        if REVERSE_RE.search(item.source_original):
            add("P3", variable_name, loc, "reverse_item", item.source_original, "确认中文已正向化且后续不再反向计分", "原题标注为反向题；按默认偏好可正向化，但需同步计分说明。", "研究者确认")
        if not item.current_chinese:
            add("P1", variable_name, loc, "missing_chinese", item.source_original, "补充中文题项或说明该行不是题项", "题项缺少当前中文文本。", "RA 可直接处理")
        source_marker = item.source_reference.lower()
        if not item.source_original and "自编" not in item.source_reference and "self-developed" not in source_marker:
            add("P2", variable_name, loc, "missing_original", item.current_chinese, "补充英文原题，或标注为自编/新增题项", "非自编量表建议保留原英文以便审查翻译与改编。", "研究者确认")

        instruction_key = (item.sheet, item.block_start_row, variable_name)
        if item.block_instruction and instruction_key not in flagged_instructions:
            if INSTRUCTION_CONTAMINATION_RE.search(item.block_instruction) and len(item.block_instruction) >= 45:
                flagged_instructions.add(instruction_key)
                priority = "P1" if "这里关注的是" in item.block_instruction or "而不是" in item.block_instruction else "P2"
                add(
                    priority,
                    variable_name,
                    f"{item.sheet}!row {item.block_start_row}",
                    "instruction_contamination",
                    item.block_instruction,
                    "缩短为中性指导语，只保留填写对象、时间窗口和反应选项；避免解释变量定义或排除相邻构念。",
                    "指导语可能提前暴露构念定义或研究者意图，影响被试填写。",
                    "研究者确认",
                )

        if item.current_chinese and REFERENT_AMBIGUITY_RE.search(item.current_chinese):
            if any(token in item.current_chinese for token in ["本部门", "部门", "我们"]):
                add(
                    "P2",
                    variable_name,
                    loc,
                    "referent_level",
                    item.current_chinese,
                    "确认理论层级后，将对象改为更明确的“我本人/我的直属领导/我所在工作小组/我直接管理的团队/正式部门”等。",
                    "题项中的群体或对象可能有多种理解，会影响领导力、团队或配对研究的 level of analysis。",
                    "研究者确认",
                )

        if "自编" in item.source_reference or "self-developed" in source_marker:
            if item.current_chinese and DOUBLE_BARRELED_RE.search(item.current_chinese) and len(item.current_chinese) >= 22:
                add(
                    "P2",
                    variable_name,
                    loc,
                    "double_barreled",
                    item.current_chinese,
                    "检查是否需要拆成单一含义题项；若保留，说明该题项只测一个核心行为或感知。",
                    "自编/高度改编题项可能同时包含多个动作、对象或条件，受访者可能只同意其中一部分。",
                    "研究者确认",
                )

        if item.source_original and item.current_chinese:
            source_lower = item.source_original.lower()
            translation_risks: list[tuple[str, str]] = []
            if "appropriately" in source_lower and "熟练" in item.current_chinese:
                translation_risks.append(("translation", "英文强调 appropriately，中文改为“熟练”可能改变强度和能力内涵。"))
            if "learn" in source_lower and "使用" in item.current_chinese and "学习" not in item.current_chinese and "掌握" not in item.current_chinese:
                translation_risks.append(("translation", "英文强调 learn/learning，中文偏向实际使用，可能改变构念含义。"))
            if "inspires me" in source_lower and "启发" in item.current_chinese:
                translation_risks.append(("translation", "英文 My job inspires me 更接近“工作激励/鼓舞我”，中文“深受启发”可能偏向认知启示。"))
            if "top managers" in source_lower and ("直属领导" in item.block_instruction or "我的直属领导" in item.current_chinese):
                translation_risks.append(("adaptation", "英文对象为 top managers，当前问卷对象为直属领导，可能改变领导层级和理论解释。"))
            if "group members" in source_lower and "我们" in item.current_chinese:
                translation_risks.append(("referent_level", "英文对象为 group members，中文“我们”可能改变回答对象和聚合层级。"))
            if "work unit" in source_lower and "部门" in item.current_chinese:
                translation_risks.append(("referent_level", "英文 work unit 需要明确是工作小组、直接管理团队还是正式部门。"))
            if "good fit" in source_lower and any(token in item.current_chinese for token in ["较差", "不匹配", "并不"]):
                translation_risks.append(("adaptation", "英文为正向 fit，中文改成不匹配/负向表达，需确认是否作为反向化或高度改编处理。"))
            for issue_type, rationale in translation_risks:
                add(
                    "P1" if issue_type in {"translation", "adaptation"} else "P2",
                    variable_name,
                    loc,
                    issue_type,
                    f"EN: {item.source_original}\nCN: {item.current_chinese}",
                    "逐条核对英文核心含义，并给出可回译的中文改法；若属于改编，补充审稿解释。",
                    rationale,
                    "研究者确认",
                )

    item_counts: dict[str, int] = {}
    respondent_by_item_var: dict[str, str] = {}
    wave_by_item_var: dict[str, str] = {}
    for item in items:
        key = variable_key(item.variable)
        if not key:
            continue
        item_counts[key] = item_counts.get(key, 0) + 1
        respondent_by_item_var.setdefault(key, item.respondent)
        wave_by_item_var.setdefault(key, item.wave)

    variable_groups: dict[str, list[VariableRecord]] = {}
    for var in variables:
        key_cn = variable_key(var.chinese_name or var.english_name)
        if key_cn:
            variable_groups.setdefault(key_cn, []).append(var)

    for key_cn, group in variable_groups.items():
        matched_keys = [
            item_key for item_key in item_counts
            if key_cn and (key_cn in item_key or item_key in key_cn)
        ]
        loc = "; ".join(f"{var.sheet}!row {var.row}" for var in group)
        label = " / ".join(v.chinese_name or v.english_name for v in group)
        if not matched_keys:
            add("P1", label, loc, "variable_missing_in_questionnaire", label, "确认该变量是否在问卷正文或另一个问卷版本中", "变量清单中有该变量，但问卷正文未能匹配到题项。", "研究者确认")
            continue
        expected = 0
        expected_known = True
        for var in group:
            try:
                expected += int(float(var.item_count))
            except ValueError:
                expected_known = False
        actual = sum(item_counts[matched_key] for matched_key in matched_keys)
        if expected_known and expected != actual:
            add("P1", label, loc, "item_count_mismatch", f"清单={expected}, 正文={actual}", "核对变量清单与问卷正文条目数", "条目数不一致会影响量表计分和问卷完整性。", "RA 可直接处理")

        variable_respondents = {respondent_alias(v.respondent) for v in group if v.respondent}
        item_respondents = {respondent_alias(respondent_by_item_var.get(k, "")) for k in matched_keys if respondent_by_item_var.get(k, "")}
        if variable_respondents and item_respondents and not item_respondents.issubset(variable_respondents):
            add("P0", label, loc, "respondent_mismatch", f"清单={','.join(sorted(variable_respondents))}, 正文={','.join(sorted(item_respondents))}", "统一填写者/评价者设置", "配对问卷中填写者不一致会直接影响数据可用性。", "研究者确认")

        variable_waves = {norm(v.wave) for v in group if v.wave}
        item_waves = {norm(wave_by_item_var.get(k, "")) for k in matched_keys if wave_by_item_var.get(k, "")}
        if variable_waves and item_waves and not item_waves.issubset(variable_waves):
            add("P1", label, loc, "wave_mismatch", f"清单={','.join(sorted(variable_waves))}, 正文={','.join(sorted(item_waves))}", "统一时间点设置", "时间点不一致会影响纵向研究设计。", "研究者确认")

    return issues


def summarize_issue(issue_type: str, variable_name: str) -> str:
    base = {
        "placeholder": "正式发放文本仍有占位符",
        "source_placeholder": "英文原量表有模板占位符，需确认中文已替换",
        "reverse_item": "原题为反向题，需确认正向化和计分说明",
        "missing_chinese": "缺少中文题项",
        "missing_original": "缺少英文原题或新增题项说明",
        "instruction_contamination": "指导语可能暴露构念定义或研究者意图",
        "referent_level": "题项 referent 或研究层级可能不清",
        "double_barreled": "自编/改编题项可能包含多个含义",
        "translation": "英文与中文可能存在语义漂移",
        "adaptation": "改编幅度可能需要审稿解释",
        "variable_missing_in_questionnaire": "变量清单中有变量，但正文未匹配到题项",
        "item_count_mismatch": "变量清单与问卷正文条目数不一致",
        "respondent_mismatch": "变量清单与正文填写者不一致",
        "wave_mismatch": "变量清单与正文时间点不一致",
    }.get(issue_type, "需要人工检查的问题")
    if variable_name:
        return f"{variable_name}：{base}"
    return base


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def human_sheet_title(name: str) -> str:
    safe = re.sub(r"[:\\/?*\[\]]", "_", name)
    return safe[:31] or "Sheet"


def as_rows(records: list[Any]) -> list[dict[str, Any]]:
    return [asdict(record) for record in records]


def add_table_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    widths: dict[str, int] | None = None,
) -> None:
    ws = workbook.create_sheet(human_sheet_title(title))
    widths = widths or {}
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    priority_fills = {
        "P0": PatternFill("solid", fgColor="F4CCCC"),
        "P1": PatternFill("solid", fgColor="FCE5CD"),
        "P2": PatternFill("solid", fgColor="FFF2CC"),
        "P3": PatternFill("solid", fgColor="D9EAD3"),
    }

    for col_idx, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(rows, start=2):
        priority = str(row.get("priority", ""))
        for col_idx, field in enumerate(fieldnames, start=1):
            value = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if priority in priority_fills:
                if field == "priority":
                    cell.fill = priority_fills[priority]
                    cell.font = Font(bold=True)
                elif title.lower() == "issues":
                    cell.fill = priority_fills[priority]
        ws.row_dimensions[row_idx].height = 72

    ws.freeze_panes = "A2"
    if fieldnames:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{max(1, len(rows) + 1)}"

    default_widths = {
        "issue_id": 12,
        "priority": 10,
        "variable_name": 26,
        "location": 18,
        "issue_type": 20,
        "issue_summary": 44,
        "current_text": 58,
        "suggested_text": 46,
        "rationale": 48,
        "reviewer_action": 18,
        "status": 14,
        "sheet": 16,
        "row": 10,
        "english_name": 26,
        "chinese_name": 28,
        "item_count": 12,
        "respondent": 12,
        "wave": 10,
        "questionnaire_id": 14,
        "variable": 34,
        "source_original": 70,
        "current_chinese": 62,
        "source_reference": 52,
        "note": 42,
        "block_instruction": 54,
    }
    default_widths.update(widths)
    for col_idx, field in enumerate(fieldnames, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = default_widths.get(field, 22)


def write_review_workbook(
    path: Path,
    summary: dict[str, Any],
    variables: list[VariableRecord],
    items: list[ItemRecord],
    issues: list[IssueRecord],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Survey Scale Review Inspection Summary"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    summary_rows = [
        ("Workbook", summary["workbook"]),
        ("Sheet count", summary["sheet_count"]),
        ("Variable count", summary["variable_count"]),
        ("Item count", summary["item_count"]),
        ("Issue count", summary["issue_count"]),
    ]
    for idx, (key, value) in enumerate(summary_rows, start=3):
        ws.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)
    start = len(summary_rows) + 5
    ws.cell(row=start, column=1, value="Priority").font = Font(bold=True)
    ws.cell(row=start, column=2, value="Count").font = Font(bold=True)
    for offset, level in enumerate(["P0", "P1", "P2", "P3"], start=1):
        ws.cell(row=start + offset, column=1, value=level)
        ws.cell(row=start + offset, column=2, value=summary["issue_counts_by_priority"].get(level, 0))
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A3"

    issue_fields = [
        "issue_id",
        "priority",
        "variable_name",
        "location",
        "issue_type",
        "issue_summary",
        "current_text",
        "suggested_text",
        "rationale",
        "reviewer_action",
        "status",
    ]
    variable_fields = list(asdict(variables[0]).keys()) if variables else [
        "sheet",
        "row",
        "english_name",
        "chinese_name",
        "item_count",
        "respondent",
        "wave",
        "questionnaire_id",
    ]
    item_fields = list(asdict(items[0]).keys()) if items else [
        "sheet",
        "row",
        "wave",
        "respondent",
        "variable",
        "source_original",
        "current_chinese",
        "source_reference",
        "note",
        "block_start_row",
        "block_instruction",
    ]

    add_table_sheet(wb, "Issues", as_rows(issues), issue_fields)
    add_table_sheet(wb, "Variables", as_rows(variables), variable_fields)
    add_table_sheet(wb, "Items", as_rows(items), item_fields)
    # Placeholder sheet for agent-generated suggestions; kept here so the workbook
    # remains the standard user-facing container after later review steps.
    suggestion_fields = [
        "variable",
        "respondent",
        "wave",
        "source_original",
        "current_chinese",
        "suggested_chinese",
        "reason",
        "priority",
    ]
    add_table_sheet(wb, "Suggestions", [], suggestion_fields)

    wb.save(path)


def write_html_table(path: Path, title: str, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    labels = {
        "issue_id": "编号",
        "priority": "优先级",
        "variable_name": "变量/量表",
        "location": "位置",
        "issue_type": "问题类型",
        "issue_summary": "问题摘要",
        "current_text": "当前文本",
        "suggested_text": "建议处理",
        "rationale": "理由",
        "reviewer_action": "处理人/动作",
        "status": "状态",
    }
    type_labels = {
        "placeholder": "占位符未替换",
        "source_placeholder": "原英文模板占位符",
        "reverse_item": "反向题正向化提醒",
        "missing_chinese": "缺少中文题项",
        "missing_original": "缺少英文原题",
        "instruction_contamination": "指导语污染风险",
        "referent_level": "Referent/层级风险",
        "double_barreled": "双重含义风险",
        "translation": "翻译等价风险",
        "adaptation": "改编解释风险",
        "variable_missing_in_questionnaire": "变量清单与正文不匹配",
        "item_count_mismatch": "条目数不一致",
        "respondent_mismatch": "填写者不一致",
        "wave_mismatch": "时间点不一致",
    }
    style = """
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;margin:28px;color:#1f2937;background:#fff}
h1{font-size:22px;margin:0 0 16px}
.hint{color:#64748b;margin:0 0 18px;font-size:13px}
.summary{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px;margin:0 0 18px}
.metric{border:1px solid #d7dee8;border-radius:8px;padding:10px;background:#f8fafc}
.metric b{display:block;font-size:20px;margin-top:4px}
.section{border:1px solid #d7dee8;border-radius:8px;padding:12px 14px;margin:0 0 18px;background:#fbfdff}
.section h2{font-size:16px;margin:0 0 8px}
.section ol,.section ul{margin:8px 0 0 22px;padding:0}
table{border-collapse:collapse;width:100%;font-size:13px;line-height:1.45}
th{position:sticky;top:0;background:#1f4e78;color:white;text-align:left}
th,td{border:1px solid #d7dee8;padding:8px;vertical-align:top}
td{white-space:pre-wrap}
.P0{background:#f4cccc}.P1{background:#fce5cd}.P2{background:#fff2cc}.P3{background:#d9ead3}
.priority{font-weight:700;text-align:center;white-space:nowrap}
.id,.status{white-space:nowrap}
"""
    counts = {
        level: sum(1 for row in rows if str(row.get("priority", "")) == level)
        for level in ["P0", "P1", "P2", "P3"]
    }
    blocker_text = "不建议发放：请先处理 P0/P1 问题。" if counts["P0"] or counts["P1"] else "未发现结构性发放阻断；仍建议人工复核翻译和改编。"
    top_rows = [row for row in rows if str(row.get("priority", "")) in {"P0", "P1"}][:3]
    top_items = "".join(
        f"<li><b>{html.escape(str(row.get('priority', '')))}</b> {html.escape(str(row.get('issue_summary', '')))} <span class='hint'>({html.escape(str(row.get('location', '')))}）</span></li>"
        for row in top_rows
    ) or "<li>当前没有 P0/P1 问题。请继续检查 P2/P3 和完整评估报告。</li>"
    parts = [
        "<!doctype html><html><head><meta charset='utf-8'>",
        f"<title>{html.escape(title)}</title><style>{style}</style></head><body>",
        f"<h1>{html.escape(title)}</h1>",
        "<p class='hint'>颜色表示优先级：P0 发放阻断，P1 高风险，P2 中风险，P3 低风险/提示。</p>",
        "<div class='section'><h2>使用边界</h2>",
        "<p>这是脚本生成的结构预检和启发式风险清单，不是完整量表审查。正式审查仍需逐题比较英文原文与中文翻译，判断改编合理性、指导语污染、自编题项质量和 referent/level-of-analysis 风险。</p>",
        "</div>",
        "<div class='summary'>",
        f"<div class='metric'>P0 发放阻断<b>{counts['P0']}</b></div>",
        f"<div class='metric'>P1 高风险<b>{counts['P1']}</b></div>",
        f"<div class='metric'>P2 中风险<b>{counts['P2']}</b></div>",
        f"<div class='metric'>P3 提示/低风险<b>{counts['P3']}</b></div>",
        "</div>",
        "<div class='section'><h2>结论</h2>",
        f"<p>{html.escape(blocker_text)}</p>",
        "<p>建议处理顺序：先修 P0，再处理 P1；RA 先处理占位符、缺失项和格式问题，研究者确认改编、构念和计分问题。</p>",
        "</div>",
        "<div class='section'><h2>前三个优先处理问题</h2><ol>",
        top_items,
        "</ol></div>",
        "<div class='section'><h2>下一步建议</h2><ul>",
        "<li>如果这是初稿：请让 RA 先处理“RA 可直接处理”的问题，再重新运行检查。</li>",
        "<li>如果准备发放：请做一次 pre-launch check，重点检查占位符、填写者、时间窗口、反应选项和配对风险。</li>",
        "<li>如果涉及英文原题：请逐行检查翻译等价，不要只看中文是否通顺。</li>",
        "<li>如果涉及高改编或自编量表：请由研究者确认改编逻辑，并考虑预测试或信效度检验。</li>",
        "</ul></div>",
        "<table><thead><tr>",
        "".join(f"<th>{html.escape(labels.get(field, field))}</th>" for field in fieldnames),
        "</tr></thead><tbody>",
    ]
    for row in rows:
        priority = str(row.get("priority", ""))
        klass = priority if priority in {"P0", "P1", "P2", "P3"} else ""
        parts.append(f"<tr class='{klass}'>")
        for field in fieldnames:
            value = str(row.get(field, ""))
            if field == "issue_type":
                value = type_labels.get(value, value)
            cell_class = "priority" if field == "priority" else "id" if field == "issue_id" else "status" if field == "status" else ""
            parts.append(f"<td class='{cell_class}'>{html.escape(value)}</td>")
        parts.append("</tr>")
    parts.extend(["</tbody></table></body></html>"])
    path.write_text("".join(parts), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--outdir", type=Path, default=Path("scale_review_inspection"))
    parser.add_argument("--xlsx", action="store_true", help="Also create a formatted Excel review workbook.")
    args = parser.parse_args()

    args.outdir.mkdir(parents=True, exist_ok=True)
    sheets = read_workbook(args.workbook)
    variables = extract_variables(sheets)
    items = extract_items(sheets)
    issues = build_issues(variables, items)

    structure = []
    for sheet_name, df in sheets.items():
        nonempty_rows = int(df.dropna(how="all").shape[0])
        nonempty_cols = int(df.dropna(axis=1, how="all").shape[1])
        structure.append(
            {
                "sheet": sheet_name,
                "shape": [int(df.shape[0]), int(df.shape[1])],
                "nonempty_rows": nonempty_rows,
                "nonempty_cols": nonempty_cols,
                "sample": sheet_sample(df),
            }
        )

    summary = {
        "workbook": str(args.workbook),
        "sheet_count": len(sheets),
        "sheets": structure,
        "variable_count": len(variables),
        "item_count": len(items),
        "issue_count": len(issues),
        "issue_counts_by_priority": {
            level: sum(1 for issue in issues if issue.priority == level)
            for level in ["P0", "P1", "P2", "P3"]
        },
    }

    (args.outdir / "inspection_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (args.outdir / "variables.json").write_text(
        json.dumps([asdict(v) for v in variables], ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (args.outdir / "items.json").write_text(
        json.dumps([asdict(i) for i in items], ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (args.outdir / "issues.json").write_text(
        json.dumps([asdict(i) for i in issues], ensure_ascii=False, indent=2), encoding="utf-8"
    )

    write_csv(args.outdir / "variables.csv", [asdict(v) for v in variables], list(asdict(variables[0]).keys()) if variables else [])
    write_csv(args.outdir / "items.csv", [asdict(i) for i in items], list(asdict(items[0]).keys()) if items else [])
    issue_fields = list(asdict(issues[0]).keys()) if issues else [
        "issue_id",
        "priority",
        "variable_name",
        "location",
        "issue_type",
        "issue_summary",
        "current_text",
        "suggested_text",
        "rationale",
        "reviewer_action",
        "status",
    ]
    write_csv(args.outdir / "issues.csv", [asdict(i) for i in issues], issue_fields)
    if args.xlsx:
        write_review_workbook(args.outdir / "scale_review_inspection.xlsx", summary, variables, items, issues)
    write_html_table(args.outdir / "issues.html", "量表问卷结构预检问题清单（非完整评审）", [asdict(i) for i in issues], issue_fields)

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
